
import PyQt4.QtGui as gui
import PyQt4.QtCore as qt
import sys
import numpy as np
from scipy.optimize import curve_fit

from openpyxl.workbook import Workbook
from openpyxl import load_workbook

from matplotlib.backends.backend_qt4agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.backends.backend_qt4agg import NavigationToolbar2QT as NavigationToolbar
from matplotlib.figure import Figure
from mpl_toolkits.mplot3d import Axes3D 
import matplotlib.pyplot as plt

import os
import shutil

header = sys.path[0]+'\\'#source path

os.chdir(header)
try:
    shutil.rmtree('./imgs')
    shutil.rmtree('./files')
    
except:
    pass

try:
    os.mkdir('./imgs')
    os.mkdir('./files')
except:
    pass

class Test(gui.QWidget):
    def __init__(self, parent = None):
        gui.QWidget.__init__(self,parent)
        '''self.setGeometry(280,80,897,620)
        self.setMaximumSize(897,620)
        self.setMinimumSize(897,620)'''
        self.setWindowTitle('Task 2 Test')
        
        self.slopes = []
        self.stds = []
        self.results = []
        print
        self.createButtons()
         
    def createButtons(self):
        self.layout = gui.QGridLayout()
        self.mainlayout = gui.QGridLayout()
        self.plotLayout = gui.QGridLayout()
        
        self.gbParam = gui.QGroupBox("PARAMETERS")
        self.ParamLayout = gui.QGridLayout()
        
        self.gbTable = gui.QGroupBox("DATA")
        self.tableLayout = gui.QGridLayout()
        
        self.gbControl = gui.QGroupBox("CONTROL")
        self.controlLayout = gui.QGridLayout()
        
        self.gbRadio = gui.QGroupBox("...")
        self.radioLayout = gui.QGridLayout()
        
        self.lblSlope = gui.QLabel('SLOPE')
        self.lblSlopeMin = gui.QLabel('min')
        self.txtSlopeMin = gui.QDoubleSpinBox()
        self.txtSlopeMin.setRange(-10,10)
        self.txtSlopeMin.setSingleStep(0.1)
        self.txtSlopeMin.setValue(-1)
        self.lblSlopeMax = gui.QLabel('max')
        self.txtSlopeMax = gui.QDoubleSpinBox()
        self.txtSlopeMax.setRange(-10,10)
        self.txtSlopeMax.setSingleStep(0.1)
        self.txtSlopeMax.setValue(1)
        self.lblSlopeNum = gui.QLabel('Samples')
        self.txtSlopeNum = gui.QSpinBox()
        self.txtSlopeNum.setRange(0,20)
        self.txtSlopeNum.setValue(10)
        
        self.lblStd = gui.QLabel('NOISE')
        self.lblStdMin = gui.QLabel('min')
        self.txtStdMin = gui.QDoubleSpinBox()
        self.txtStdMin.setRange(0,100)
        self.txtStdMin.setSingleStep(0.01)
        self.txtStdMin.setValue(0)
        self.lblStdMax = gui.QLabel('max')
        self.txtStdMax = gui.QDoubleSpinBox()
        self.txtStdMax.setRange(0,100)
        self.txtStdMax.setSingleStep(0.01)
        self.txtStdMax.setValue(10)
        self.lblStdNum = gui.QLabel('Samples')
        self.txtStdNum = gui.QSpinBox()
        self.txtStdNum.setRange(0,20)
        self.txtStdNum.setValue(10)
        
        self.lblXscale = gui.QLabel('X-Scaling Factor')
        self.txtXscale = gui.QDoubleSpinBox()
        self.txtXscale.setRange(-100,100)
        self.txtXscale.setSingleStep(0.01)
        self.txtXscale.setValue(0.1)
        
        self.lblSamples = gui.QLabel('Data Samples')
        self.txtSamples = gui.QSpinBox()
        self.txtSamples.setRange(1024,16384)
        self.txtSamples.setValue(1024)
        
        self.btn = gui.QPushButton('Run Test')
        self.btn.clicked.connect(self.run_test)
        self.cbFiles = gui.QCheckBox('Save Files')
        self.cbPlots = gui.QCheckBox('Save Plots')
        
        self.rb3d = gui.QRadioButton('3D Plot')
        self.rb3d.clicked.connect(self.plot)
        self.rb2d = gui.QRadioButton('2D Plot')
        self.rb2d.setChecked(True)
        self.rb2d.clicked.connect(self.plot)
        
        self.tblData = gui.QTableWidget()
        self.tblData.setRowCount(5)
        self.tblData.setColumnCount(2)
        self.tblData.setHorizontalHeaderLabels(qt.QString("X-Data,Y-Data").split(','))
        self.tblData.show()
        
        self.figure = Figure()
        self.canvas = FigureCanvas(self.figure)
        self.toolbar = NavigationToolbar(self.canvas, self)
        
        hbox = gui.QHBoxLayout()
        hbox.addWidget(self.lblSamples)
        hbox.addWidget(self.txtSamples)
        hbox.addWidget(self.lblXscale)
        hbox.addWidget(self.txtXscale)
        self.ParamLayout.addLayout(hbox,0,0)
        
        vbox0 = gui.QVBoxLayout()
        vbox0.addWidget(self.lblSlope)
        vbox1 = gui.QVBoxLayout()
        vbox1.addWidget(self.lblSlopeMin)
        vbox1.addWidget(self.txtSlopeMin)
        vbox2 = gui.QVBoxLayout()
        vbox2.addWidget(self.lblSlopeMax)
        vbox2.addWidget(self.txtSlopeMax)
        vbox3 = gui.QVBoxLayout()
        vbox3.addWidget(self.lblSlopeNum)
        vbox3.addWidget(self.txtSlopeNum)
        hbox = gui.QHBoxLayout()
        hbox.addLayout(vbox0)
        hbox.addLayout(vbox1)
        hbox.addLayout(vbox2)
        hbox.addLayout(vbox3)
        self.ParamLayout.addLayout(hbox,1,0)
        
        vbox0 = gui.QVBoxLayout()
        vbox0.addWidget(self.lblStd)
        vbox1 = gui.QVBoxLayout()
        vbox1.addWidget(self.lblStdMin)
        vbox1.addWidget(self.txtStdMin)
        vbox2 = gui.QVBoxLayout()
        vbox2.addWidget(self.lblStdMax)
        vbox2.addWidget(self.txtStdMax)
        vbox3 = gui.QVBoxLayout()
        vbox3.addWidget(self.lblStdNum)
        vbox3.addWidget(self.txtStdNum)
        hbox = gui.QHBoxLayout()
        hbox.addLayout(vbox0)
        hbox.addLayout(vbox1)
        hbox.addLayout(vbox2)
        hbox.addLayout(vbox3)
        self.ParamLayout.addLayout(hbox,2,0)
        self.gbParam.setLayout(self.ParamLayout)
        self.layout.addWidget(self.gbParam,1,0)
        
        hbox = gui.QHBoxLayout()
        hbox.addWidget(self.tblData)
        vbox = gui.QVBoxLayout()
        vbox.addLayout(hbox)
        self.tableLayout.addLayout(vbox,0,0)
        self.gbTable.setLayout(self.tableLayout)
        
        vbox = gui.QVBoxLayout()
        vbox.addWidget(self.rb2d)
        vbox.addWidget(self.rb3d)
        hbox = gui.QHBoxLayout()
        hbox.addLayout(vbox)
        self.radioLayout.addLayout(hbox,0,0)
        self.gbRadio.setLayout(self.radioLayout)
        
        vbox = gui.QVBoxLayout()
        vbox.addWidget(self.btn)
        vbox.addWidget(self.cbFiles)
        vbox.addWidget(self.cbPlots)
        vbox.addWidget(self.gbRadio)
        hbox = gui.QHBoxLayout()
        hbox.addLayout(vbox)
        self.controlLayout.addLayout(hbox,0,0)
        self.gbControl.setLayout(self.controlLayout)
        
        
        
        hbox = gui.QHBoxLayout()
        hbox.addWidget(self.gbTable)
        hbox.addWidget(self.gbControl)
        self.layout.addLayout(hbox,2,0)
        
        
        vbox = gui.QVBoxLayout()
        vbox.addWidget(self.toolbar)
        vbox.addWidget(self.canvas)
        hbox = gui.QHBoxLayout()
        hbox.addLayout(vbox)
        self.plotLayout.addLayout(hbox,0,0)
        
        
        hbox = gui.QHBoxLayout()
        hbox.addLayout(self.plotLayout)
        hbox.addLayout(self.layout)
        
        self.mainlayout.addLayout(hbox, 0, 0)
        
        self.setLayout(self.mainlayout)
    
    def error(self, msg):
        gui.QMessageBox.warning(self, 'Error', msg)

    def info(self, msg):
        gui.QMessageBox.information(self, 'Info', msg)
    
    def run_test(self):
        self.results = []
        self.slopes = np.linspace(self.txtSlopeMin.value(), self.txtSlopeMax.value(), self.txtSlopeNum.value())
        self.stds = np.linspace(self.txtStdMin.value(), self.txtStdMax.value(), self.txtStdNum.value())
        
        for i in range(len(self.slopes)):
            for j in range(len(self.stds)):
                title = 'slope_'+str(self.slopes[i])+' std_'+str(self.stds[j])
                self.results.append(Data(title, m = self.slopes[i], std = self.stds[j], num_samples = self.txtSamples.value(), x_factor = self.txtXscale.value(), files = self.cbFiles.isChecked(), plots = self.cbPlots.isChecked()))
        
        self.res = [np.round(i.data['gradient']-i.data['grad_pred'], 3) for i in self.results]
        self.grad_pred = [np.round(i.data['grad_pred'],2) for i in self.results]
        self.res = np.reshape(self.res,(len(self.slopes), len(self.stds)))
        
        self.tblData.clear()
        
        self.tblData.setColumnCount(len(self.stds))
        self.tblData.setRowCount(len(self.slopes))
        
        HLabels = ''
        for i in self.slopes:
            HLabels += str(np.round(i,2))+','
        HLabels.rstrip()
        
        VLabels = ''
        for i in self.stds:
            VLabels += str(np.round(i,2))+','
        VLabels.rstrip()
        
        self.tblData.setHorizontalHeaderLabels(qt.QString(VLabels).split(','))
        self.tblData.setVerticalHeaderLabels(qt.QString(HLabels).split(','))
        
        for i in range(len(self.slopes)):
            for j in range(len(self.stds)):
                self.tblData.setItem(i,j,gui.QTableWidgetItem(str(self.res[i,j]))) 
        
        self.plot()
    
    def plot(self):
        if self.rb2d.isChecked():
            self.plot2D()
            
        elif self.rb3d.isChecked():
            self.plot3D()
    
    def plot3D(self):
        try:
            
            '''plt.imshow(res)
            plt.colorbar()
            plt.grid()
            plt.show()'''
            
            # Data generation
            alpha = self.slopes
            t = self.stds
            T, A = np.meshgrid(t, alpha)
            data = np.abs(self.res)#self.res+(np.abs(np.min(self.res)))
            
            # Plotting
            
            
            self.figure.clear()
            ax = self.figure.gca(projection = '3d')
            ax.clear()
            
            Xi = T.flatten()
            Yi = A.flatten()
            Zi = np.zeros(data.size)
            
            dx = .25 * np.ones(data.size)
            dy = .25 * np.ones(data.size)
            dz = data.flatten()
            
            ax.bar3d(Xi, Yi, Zi, dx, dy, dz, color = 'w')
            
            ax.set_zlabel('Errors')
            ax.set_ylabel('Slope')
            ax.set_xlabel('Standard Deviation of Noise')
            ax.set_title('Slope Prediction Errors')
            
            
            #ax.legend(['prediction','data'],loc='upper left',prop={'size':12})
            
            self.canvas.draw()
            
        except Exception as e:
            self.error(str(e))
    
    def plot2D(self):
        
        try:
            print np.average(np.abs(self.res),0)
            #TODO: add radio button to toggle between plots 2d average and 3d full
            
            self.figure.clear()
            ax = self.figure.add_subplot(111)
            ax.clear()
            
            #ax.bar(self.stds, np.average(np.abs(self.res),0))
            ax.plot(self.stds, np.average(np.abs(self.res),0), '-o')
            
            ax.set_xlabel('Standard Deviation of Noise')
            ax.set_ylabel('Average Error')
            ax.set_title('Average Error in Slope Prediction')
            
            
            self.canvas.draw()
            
        except Exception as e:
            self.error(str(e))
        
class Data(object):
    def __init__(self, name, m = 5, std = 0.4, num_samples = 100, x_factor = 1, files = False, plots = False):
        self.name = name
        self.m = m
        self.std = std
        self.num = num_samples
        self.x_factor = x_factor
        
        noise = np.random.normal(0, self.std, size=self.num)
        #print self.std
        #print noise[[0,1,2,3,4,5]]
        self.x_data = (np.array(range(self.num)))*self.x_factor
        Y = (self.m*self.x_data) + noise
        self.y_data = self.preprocess_list_line(Y.reshape(self.num,1))
        self.gradient = None
        self.intercept = None
        self.y_predict = None
        self.figure = None
        
        self.evaluate()
        
        self.error = self.m - self.gradient
        if plots:
            self.save_plot()
        if files:
            self.write_xls()
        
        self.data = dict()
        self.data['name'] = name
        self.data['gradient'] = self.m
        self.data['y_data'] = self.y_data
        self.data['num'] = self.num
        self.data['std'] = self.std
        self.data['y_pred'] = self.y_predict
        self.data['grad_pred'] = self.gradient
        self.data['error'] = self.error
        
    def __str__(self):
        return str(self.data)
    
    def evaluate(self):
        self.gradient, self.intercept = curve_fit(self.line, self.x_data, self.y_data)[0]
        self.y_predict = self.line(self.x_data, self.gradient, self.intercept)
        
    def line(self, x, M, C):
        """line equation for fitting"""
        return M*x + C
    
    def preprocess_list_line(self, data):
        """return a list of Y data from generated data"""
        bn = []
        [bn.append(float(i[0])) for i in data]
        
        return bn
    
    def write_xls(self):
        book = Workbook()
        sheet = book.active
        for i in range(self.num):
            sheet['A'+str(i+1)] = self.y_data[i]
        book.save(header+'files//'+str(self.name)+'.xlsx')
    
    def save_plot(self):
        self.figure,ax = plt.subplots(1,1)
        ax.clear()
        ax.hold(True)
        
        ax.scatter(self.x_data, self.y_data)
        ax.set_xlabel('x_data')
        ax.set_ylabel('y_data')
        ax.set_title('Slope Prediction '+ self.name)
        
        ax.plot(self.x_data,self.y_predict,'r-', linewidth=2)
        ax.legend(['prediction','data'],loc='upper left',prop={'size':12})
        
        self.figure.savefig(header+'imgs//'+str(self.name)+'.png')
        
class Window(gui.QWidget):
    def __init__(self, parent = None):
        gui.QWidget.__init__(self,parent)
        
        self.test = Test()
        
        self.setGeometry(280,80,897,620)
        self.setMaximumSize(897,620)
        self.setMinimumSize(897,620)
        self.setWindowTitle('Automation Lab - Task 2')
        
        self.w = gui.QWidget()
        self.w.resize(320, 240)
        self.w.setWindowTitle("Select Line Data File")
        
        self.gradient = None
        self.intercept = None
        self.y_predict = None
        
        self.createButtons()
         
    def select_file(self):
        filename = gui.QFileDialog.getOpenFileName(self.w, 'Open File', header)
        self.txtPath.setText(filename)
        
    def createButtons(self):
        self.layout = gui.QGridLayout()
        
        self.gbPlot = gui.QGroupBox("PLOTS")
        self.plotLayout = gui.QGridLayout()
        
        self.gbTable = gui.QGroupBox("DATA")
        self.tableLayout = gui.QGridLayout()
        
        self.gbRange = gui.QGroupBox("...")
        self.rangeLayout = gui.QGridLayout()
        
        self.lblPath = gui.QLabel("Path")
        self.txtPath = gui.QLineEdit(header+"line.xlsx")
        self.btnSelect = gui.QPushButton("File")
        self.btnSelect.clicked.connect(self.select_file)
        
        '''self.lblXmin = gui.QLabel('X-min')
        self.txtXmin = gui.QDoubleSpinBox()
        self.txtXmin.setRange(-100000,100000)
        self.txtXmin.setSingleStep(0.01)
        self.txtXmin.setValue(0)
        self.lblXmax = gui.QLabel('X-max')
        self.txtXmax = gui.QDoubleSpinBox()
        self.txtXmax.setRange(-100000,100000)
        self.txtXmax.setSingleStep(0.01)
        self.txtXmax.setValue(100)'''
        self.lblXscale = gui.QLabel('x-Scaling Factor')
        self.txtXscale = gui.QDoubleSpinBox()
        self.txtXscale.setRange(-100000,100000)
        self.txtXscale.setSingleStep(0.01)
        self.txtXscale.setValue(1)
        
        self.lblSlope = gui.QLabel('Estimated Slope')
        self.txtSlope = gui.QLabel()
        self.bluepalette = gui.QPalette()
        c=gui.QColor(50,100,200,255)
        self.bluepalette.setColor(gui.QPalette.Foreground,c)
        self.txtSlope.setFont(gui.QFont('SansSerif', 18))
        self.txtSlope.setPalette(self.bluepalette)
        
        self.btnTest = gui.QPushButton('Run Test')
        self.btnTest.clicked.connect(self.runtest)
        
        self.figure = Figure()
        self.canvas = FigureCanvas(self.figure)
        self.toolbar = NavigationToolbar(self.canvas, self)
        self.btn = gui.QPushButton('Plot')
        self.btn.clicked.connect(self.plot)
        
        self.tblData = gui.QTableWidget()
        self.tblData.setRowCount(5)
        self.tblData.setColumnCount(2)
        self.tblData.setHorizontalHeaderLabels(qt.QString("X-Data,Y-Data").split(','))
        self.tblData.show()
        
        hbox = gui.QHBoxLayout()
        hbox.addWidget(self.lblPath)
        hbox.addWidget(self.txtPath)
        hbox.addWidget(self.btnSelect)
        vbox = gui.QVBoxLayout()
        vbox.addLayout(hbox)
        self.layout.addLayout(vbox,0,0)
        
        '''vbox1 = gui.QVBoxLayout()
        vbox1.addWidget(self.lblXmin)
        vbox1.addWidget(self.txtXmin)
        vbox2 = gui.QVBoxLayout()
        vbox2.addWidget(self.lblXmax)
        vbox2.addWidget(self.txtXmax)
        vbox3 = gui.QVBoxLayout()
        vbox3.addWidget(self.lblSlope)
        vbox3.addWidget(self.txtSlope)
        hbox = gui.QHBoxLayout()
        hbox.addLayout(vbox1)
        hbox.addLayout(vbox2)
        hbox.addLayout(vbox3)
        self.rangeLayout.addLayout(hbox,0,0)
        self.gbRange.setLayout(self.rangeLayout)
        self.layout.addWidget(self.gbRange,1,0)'''
        
        vbox1 = gui.QVBoxLayout()
        vbox1.addWidget(self.lblXscale)
        vbox1.addWidget(self.txtXscale)
        
        vbox3 = gui.QVBoxLayout()
        vbox3.addWidget(self.lblSlope)
        vbox3.addWidget(self.txtSlope)
        hbox = gui.QHBoxLayout()
        hbox.addLayout(vbox1)
        hbox.addLayout(vbox3)
        self.rangeLayout.addLayout(hbox,0,0)
        self.gbRange.setLayout(self.rangeLayout)
        self.layout.addWidget(self.gbRange,1,0)
        
        hbox = gui.QVBoxLayout()
        hbox.addWidget(self.toolbar)
        hbox.addWidget(self.canvas)
        
        box = gui.QHBoxLayout()
        box.addWidget(self.btn)
        box.addWidget(self.btnTest)
        hbox.addLayout(box)
        vbox = gui.QHBoxLayout()
        vbox.addLayout(hbox)
        self.plotLayout.addLayout(vbox,0,0)
        self.gbPlot.setLayout(self.plotLayout)
        
        hbox = gui.QHBoxLayout()
        hbox.addWidget(self.tblData)
        vbox = gui.QVBoxLayout()
        vbox.addLayout(hbox)
        self.tableLayout.addLayout(vbox,0,0)
        self.gbTable.setLayout(self.tableLayout)
        
        hbox = gui.QHBoxLayout()
        hbox.addWidget(self.gbPlot)
        hbox.addWidget(self.gbTable)
        self.layout.addLayout(hbox,2,0)
        
        self.setLayout(self.layout)
    
    def error(self, msg):
        gui.QMessageBox.warning(self, 'Error', msg)

    def info(self, msg):
        gui.QMessageBox.information(self, 'Info', msg)
    
    def read_file(self):
        """read a file form source folder"""
        try:
            f = open(self.txtPath.text(), 'r')
            raw =  str(f.read())
            f.close()
            return raw
        except Exception as e:
            self.error("Please Select File Path To Read")
            return None
    
    def write_xls(self, data, name):
        book = Workbook()
        sheet = book.active
        for i in range(len(data)):
            sheet['A'+str(i+1)] = data[i]
        book.save(name)
        
    def read_xls(self, name):
        """read a excel file form path"""
        book = load_workbook(name)
        sheet = book.active
        data = []
        ind = 1
        while True:
            y = sheet['A' +str(ind)].value
            ind+=1
            if y == None:
                break
            else:
                data.append(y)
        return data
    
    def preprocess(self, data):
        """return a list of Y data from read file"""
        data = str(data)
        res = data.replace('[','')
        res = res.replace(']', '')
        res = res.replace(' ', '')
        res = res.split('\n')
        bn = []
        [bn.append(float(i)) for i in res]
        return bn
    
    def line(self, x, M, C):
        """line equation for fitting"""
        return M*x + C
    
    def plot(self):
        
        #file_ = self.read_file()
        #self.y_data = self.preprocess(file_)
        self.y_data = self.read_xls(str(self.txtPath.text()))
        n = len(self.y_data)
        #self.x_data = np.linspace(float(self.txtXmin.value()), float(self.txtXmax.value()), n)
        self.x_data = (np.array(range(n)))*self.txtXscale.value()
        
        self.tblData.setRowCount(n)
        [self.tblData.setItem(i,0,gui.QTableWidgetItem(str(self.x_data[i]))) for i in range(n)]
        [self.tblData.setItem(i,1,gui.QTableWidgetItem(str(self.y_data[i]))) for i in range(n)]
        
        self.figure.clear()
        ax = self.figure.add_subplot(111)
        ax.clear()
        ax.hold(True)
        
        ax.scatter(self.x_data, self.y_data)
        ax.set_xlabel('x_data')
        ax.set_ylabel('y_data')
        ax.set_title('Slope Prediction')
        
        
        self.gradient, self.intercept = curve_fit(self.line, self.x_data, self.y_data)[0]
        #print 'curve_fit: '+str(self.gradient) +' '+str(self.intercept)
        self.txtSlope.setText(str(self.gradient))
        
        self.y_predict = self.line(self.x_data, self.gradient, self.intercept)
        ax.plot(self.x_data,self.y_predict,'r-', linewidth=2)
        
        ax.legend(['prediction','data'],loc='upper left',prop={'size':12})
        
        self.canvas.draw()
    
    def runtest(self):
        self.test.show()
        
if __name__ == '__main__':
    app=gui.QApplication(sys.argv)
    
    win=Window()
    #win=Test()
    win.show()
    sys.exit(app.exec_())

